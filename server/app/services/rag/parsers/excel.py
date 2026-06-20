"""Excel 文本提取：.xlsx（openpyxl）、.xls（xlrd）。"""

from pathlib import Path

from app.services.rag.errors import IngestError
from app.services.rag.parsers._utils import require_non_empty


def _format_sheet_block(*, title: str | None, lines: list[str]) -> str:
    block = "\n".join(lines)
    if title:
        return f"【{title}】\n{block}"
    return block


def _row_to_line(values: tuple[object, ...]) -> str | None:
    cells = [str(value).strip() for value in values if value is not None and str(value).strip()]
    if not cells:
        return None
    return "\t".join(cells)


def parse_xlsx_file(file_path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise IngestError("XLSX 解析依赖未安装，请执行 pip install openpyxl") from exc

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as exc:
        raise IngestError(f"无法读取 XLSX: {exc}") from exc

    parts: list[str] = []
    try:
        multi_sheet = len(workbook.worksheets) > 1
        for worksheet in workbook.worksheets:
            lines: list[str] = []
            for row in worksheet.iter_rows(values_only=True):
                line = _row_to_line(row)
                if line:
                    lines.append(line)
            if lines:
                title = worksheet.title if multi_sheet else None
                parts.append(_format_sheet_block(title=title, lines=lines))
    finally:
        workbook.close()

    return require_non_empty("\n\n".join(parts), empty_message="XLSX 未提取到文本")


def parse_xls_file(file_path: Path) -> str:
    try:
        import xlrd
    except ImportError as exc:
        raise IngestError("XLS 解析依赖未安装，请执行 pip install xlrd") from exc

    try:
        workbook = xlrd.open_workbook(file_path)
    except Exception as exc:
        raise IngestError(f"无法读取 XLS: {exc}") from exc

    parts: list[str] = []
    multi_sheet = workbook.nsheets > 1
    for sheet in workbook.sheets():
        lines: list[str] = []
        for row_idx in range(sheet.nrows):
            row_values = tuple(sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols))
            line = _row_to_line(row_values)
            if line:
                lines.append(line)
        if lines:
            title = sheet.name if multi_sheet else None
            parts.append(_format_sheet_block(title=title, lines=lines))

    return require_non_empty("\n\n".join(parts), empty_message="XLS 未提取到文本")
